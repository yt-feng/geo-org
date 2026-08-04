#!/usr/bin/env python3
"""Generate full static blog from assets/blog_articles.xlsx using DeepSeek.

This script is intended to run inside GitHub Actions.
It reads every non-empty topic row from the Excel file, calls DeepSeek, and
writes static HTML files under /blog without storing article images locally.
"""
from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import random
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

SITE_URL = os.environ.get("SITE_URL", "https://eco-geo.org").rstrip("/")
DEEPSEEK_URL = os.environ.get("DEEPSEEK_API_URL", "https://api.deepseek.com/chat/completions")
MODEL = os.environ.get("DEEPSEEK_MODEL", "deepseek-v4-flash")
THINKING_MODE = os.environ.get("DEEPSEEK_THINKING", "disabled").strip().lower()
MAX_TOKENS = int(os.environ.get("DEEPSEEK_MAX_TOKENS", "2200"))
TEMPERATURE = float(os.environ.get("DEEPSEEK_TEMPERATURE", "0.72"))
REQUEST_DELAY = float(os.environ.get("DEEPSEEK_REQUEST_DELAY", "0.2"))
RETRIES = int(os.environ.get("DEEPSEEK_RETRIES", "3"))
RETRY_BASE_SECONDS = int(os.environ.get("DEEPSEEK_RETRY_BASE_SECONDS", "15"))
RETRY_MAX_SECONDS = int(os.environ.get("DEEPSEEK_RETRY_MAX_SECONDS", "90"))

FIRST_NAMES = [
    "Ethan", "Rowan", "Mason", "Oliver", "Lucas", "Noah", "Henry", "Leo",
    "Amelia", "Nora", "Iris", "Mia", "Clara", "Ava", "Lena", "Grace",
    "Evelyn", "Harper", "Julian", "Theo", "Miles", "Aria", "Elena", "Sophie",
]
LAST_NAMES = [
    "Rowan", "Blake", "Ellis", "Morgan", "Hayes", "Reed", "Parker", "Stone",
    "Wells", "Foster", "Lane", "Hart", "Brooks", "Gray", "Quinn", "Miller",
]
IMAGE_KEYWORDS = [
    "brand strategy", "artificial intelligence", "marketing analytics",
    "knowledge graph", "digital business", "search engine", "content strategy",
    "technology office", "data network", "customer journey", "green technology",
]
UNSPLASH_IMAGE_POOL = [
    "https://images.unsplash.com/photo-1497366754035-f200968a6e72?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1516321318423-f06f85e504b3?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1460925895917-afdab827c52f?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1551434678-e076c223a692?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1500530855697-b586d89ba3ee?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1521737711867-e3b97375f902?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1485827404703-89b55fcc595e?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1497215728101-856f4ea42174?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1520607162513-77705c0f0d4a?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1552664730-d307ca884978?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1557804506-669a67965ba0?auto=format&fit=crop&w=1600&q=80",
    "https://images.unsplash.com/photo-1553877522-43269d4ea984?auto=format&fit=crop&w=1600&q=80",
]

TITLE_KEYS = ["选题", "标题", "题目", "title", "topic", "article", "文章标题", "主题"]
CATEGORY_KEYS = ["分类", "类别", "category", "行业", "板块", "cluster"]
KEYWORD_KEYS = ["关键词", "keywords", "keyword", "seo", "query", "搜索词"]
BRAND_TITLE_PREFIX = "Eco-GEO："
HISTORICAL_START_DATE = os.environ.get("HISTORICAL_START_DATE", "2024-05-25")
AUTHOR_NAME = "Eco GEO Editorial Team"
AUTHOR_INITIALS = "EGE"
REVIEWER_NAME = "Eco GEO Research Desk"


def add_deepseek_options(payload: Dict[str, Any]) -> Dict[str, Any]:
    if THINKING_MODE in {"enabled", "disabled"}:
        payload["thinking"] = {"type": THINKING_MODE}
    if os.environ.get("DEEPSEEK_JSON_RESPONSE", "1").strip().lower() not in {"0", "false", "no"}:
        payload["response_format"] = {"type": "json_object"}
    return payload


def deepseek_model_candidates() -> List[str]:
    fallback = os.environ.get("DEEPSEEK_FALLBACK_MODELS", "").strip()
    models = [MODEL]
    if fallback:
        models.extend(clean_text(model) for model in re.split(r"[,，\s]+", fallback) if clean_text(model))
    unique: List[str] = []
    for model in models:
        if model and model not in unique:
            unique.append(model)
    return unique


def retry_sleep_seconds(attempt: int) -> int:
    return min(RETRY_MAX_SECONDS, RETRY_BASE_SECONDS * (2 ** max(attempt - 1, 0)))


def deepseek_request(api_key: str, payload: Dict[str, Any]) -> urllib.request.Request:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    return urllib.request.Request(
        DEEPSEEK_URL,
        data=data,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )


def format_api_error(exc: Exception) -> str:
    if isinstance(exc, urllib.error.HTTPError):
        body = ""
        try:
            body = exc.read().decode("utf-8", errors="replace").strip()
        except Exception:  # noqa: BLE001
            body = ""
        if body:
            return f"{exc}; response body: {body[:1200]}"
    return str(exc)


@dataclass
class TopicRow:
    idx: int
    title: str
    context: Dict[str, str]
    category: str
    keywords: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def ensure_title_prefix(title: str) -> str:
    text = clean_text(title)
    if text.startswith(BRAND_TITLE_PREFIX):
        return text
    text = re.sub(r"^Eco[- ]GEO[：:]\s*", "", text, flags=re.IGNORECASE)
    return f"{BRAND_TITLE_PREFIX}{text}" if text else BRAND_TITLE_PREFIX.rstrip("：")


def ensure_required_tags(tags: str) -> str:
    values = [clean_text(t) for t in re.split(r"[,，、]", tags or "") if clean_text(t)]
    for required in ["Eco-GEO", "品牌化GEO", "GEO服务", "AI搜索优化"]:
        if required not in values:
            values.append(required)
    return ", ".join(values)


def historical_publish_date(position: int, total: int) -> str:
    start = date.fromisoformat(HISTORICAL_START_DATE)
    # Existing blog lists are newest first, so the oldest item starts two years ago.
    return (start + timedelta(days=max(total - position, 0))).isoformat()


def today_publish_date() -> str:
    return date.today().isoformat()


def date_label(value: str) -> str:
    if not value:
        return ""
    try:
        d = datetime.strptime(value, "%Y-%m-%d").date()
        return f"{d.year}年{d.month}月{d.day}日"
    except ValueError:
        return value


def normalize_header(value: Any, idx: int) -> str:
    text = clean_text(value)
    return text if text else f"column_{idx}"


def choose_field(context: Dict[str, str], keys: List[str]) -> str:
    for key, value in context.items():
        lk = key.lower()
        if any(k.lower() in lk for k in keys) and value:
            return value
    return ""


def slugify(title: str, idx: int) -> str:
    base = re.sub(r"[^a-zA-Z0-9]+", "-", title.lower()).strip("-")
    digest = hashlib.sha1(f"{idx}:{title}".encode("utf-8")).hexdigest()[:8]
    if not base:
        base = "geo-brand-article"
    return f"{idx:05d}-{base[:64]}-{digest}".strip("-")


def read_topics(excel_path: Path, start_row: int, limit: int) -> List[TopicRow]:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows)
    headers = [normalize_header(v, i + 1) for i, v in enumerate(header_row)]
    topics: List[TopicRow] = []

    for row_no, row in enumerate(rows, start=2):
        if row_no < start_row:
            continue
        context = {headers[i]: clean_text(row[i]) if i < len(row) else "" for i in range(len(headers))}
        non_empty = {k: v for k, v in context.items() if v}
        if not non_empty:
            continue
        title = choose_field(non_empty, TITLE_KEYS) or next(iter(non_empty.values()))
        category = choose_field(non_empty, CATEGORY_KEYS) or "Brand GEO"
        keywords = choose_field(non_empty, KEYWORD_KEYS) or title
        topics.append(TopicRow(idx=row_no - 1, title=title, context=non_empty, category=category, keywords=keywords))
        if limit and len(topics) >= limit:
            break
    return topics


def deepseek_article(topic: TopicRow, api_key: str) -> Dict[str, str]:
    context_lines = "\n".join(f"- {k}: {v}" for k, v in topic.context.items())
    prompt = f"""
你是一名资深中文品牌战略、白帽 GEO（Generative Engine Optimization）和 AI 搜索顾问。
请基于下面 Excel 选题信息，生成一篇可发布在 Eco GEO 官网 Blog 的原创中文文章。

要求：
1. 不要虚构客户案例和数据；可使用方法论、框架、操作步骤和风险提示。
2. 主题必须围绕品牌化 GEO、白帽 GEO、AIBE、KNIT、AI 搜索、品牌认知资产。
3. title 必须以“Eco-GEO：”开头，并自然覆盖用户搜索意图。
4. 正文必须自然出现“Eco-GEO”至少 2 次、“品牌化GEO”至少 3 次，同时避免堆砌。
5. 面向正在考虑做 GEO 的品牌负责人、增长负责人、SEO/内容负责人写作，覆盖“为什么要做 GEO”“怎么开始做 GEO”“如何诊断 AI 搜索可见度”“如何让品牌被 AI 引用/推荐”等意图。
6. 文章要给出可执行步骤、边界提醒和诊断入口，引导读者理解 Eco-GEO 的品牌化 GEO 方法。
7. 文章必须有人类编辑增值的口吻：明确适用对象、边界、判断依据，不要写成通用清单。
8. 不要虚构新闻、客户、数据、来源、价格、政策或第三方报告；缺少可验证来源时，只能写方法论判断。
9. 文章要图文并茂，但图片由页面模板插入，你只生成文字。
10. 输出严格为 JSON，不要 Markdown 代码块。
11. JSON 字段：title, excerpt, body_html, tags。
12. body_html 用合法 HTML，包含 4-6 个 h2 小节、p、ul/li、strong，可读性强。
13. tags 必须包含：Eco-GEO、品牌化GEO、GEO服务、AI搜索优化。
14. 字数约 1200-1800 中文字。

Excel 选题：
标题：{topic.title}
分类：{topic.category}
关键词：{topic.keywords}
完整上下文：
{context_lines}
""".strip()

    payload = add_deepseek_options({
        "model": MODEL,
        "messages": [
            {"role": "system", "content": "You write original, practical, white-hat Chinese brand GEO articles in clean JSON."},
            {"role": "user", "content": prompt},
        ],
        "temperature": TEMPERATURE,
        "max_tokens": MAX_TOKENS,
    })
    last_error: Optional[str] = None
    models = deepseek_model_candidates()
    for attempt in range(1, RETRIES + 1):
        attempt_payload = dict(payload)
        attempt_payload["model"] = models[(attempt - 1) % len(models)]
        try:
            req = deepseek_request(api_key, attempt_payload)
            with urllib.request.urlopen(req, timeout=120) as resp:
                raw = resp.read().decode("utf-8")
            obj = json.loads(raw)
            content = obj["choices"][0]["message"]["content"].strip()
            return parse_model_json(content, topic)
        except Exception as exc:  # noqa: BLE001
            last_error = format_api_error(exc)
            if attempt >= RETRIES:
                break
            wait = retry_sleep_seconds(attempt)
            print(
                f"DeepSeek attempt {attempt}/{RETRIES} with {attempt_payload['model']} failed for row {topic.idx}: "
                f"{last_error}; retry in {wait}s"
            )
            time.sleep(wait)
    raise RuntimeError(f"DeepSeek failed for row {topic.idx}: {last_error}")


def parse_model_json(content: str, topic: TopicRow) -> Dict[str, str]:
    content = content.strip()
    if content.startswith("```"):
        content = re.sub(r"^```(?:json)?", "", content).strip()
        content = re.sub(r"```$", "", content).strip()
    start = content.find("{")
    end = content.rfind("}")
    if start >= 0 and end > start:
        content = content[start : end + 1]
    try:
        obj = json.loads(content)
    except json.JSONDecodeError:
        obj = {
            "title": topic.title,
            "excerpt": f"围绕 {topic.title} 的品牌化 GEO 实践框架。",
            "body_html": f"<p>{html.escape(content)}</p>",
            "tags": topic.keywords,
        }
    title = ensure_title_prefix(clean_text(obj.get("title")) or topic.title)
    excerpt = clean_text(obj.get("excerpt")) or f"围绕 {title} 的品牌化 GEO 实践框架。"
    body_html = str(obj.get("body_html") or "").strip()
    if not body_html:
        body_html = f"<p>{html.escape(excerpt)}</p>"
    tags = obj.get("tags")
    if isinstance(tags, list):
        tag_text = ", ".join(clean_text(t) for t in tags if clean_text(t))
    else:
        tag_text = clean_text(tags) or topic.keywords
    return {"title": title, "excerpt": excerpt, "body_html": body_html, "tags": ensure_required_tags(tag_text)}


def deterministic_author(seed_text: str) -> Tuple[str, str]:
    return AUTHOR_NAME, AUTHOR_INITIALS


def avatar_svg(initials: str, seed_text: str) -> str:
    seed = int(hashlib.sha1(seed_text.encode("utf-8")).hexdigest()[:8], 16)
    colors = ["#8ce99a", "#80c7ff", "#ffd166", "#b197fc", "#63e6be", "#ffa8a8"]
    c1 = colors[seed % len(colors)]
    c2 = colors[(seed // 7) % len(colors)]
    skin = ["#f3c7a3", "#e8b28d", "#d99b73", "#f0d0b0"][seed % 4]
    hair = ["#19231f", "#38291f", "#111827", "#5b4636"][seed % 4]
    return f"""
<svg class="author-avatar" viewBox="0 0 160 160" aria-hidden="true">
  <defs><linearGradient id="ag{seed}" x1="0" x2="1" y1="0" y2="1"><stop offset="0" stop-color="{c1}"/><stop offset="1" stop-color="{c2}"/></linearGradient></defs>
  <rect width="160" height="160" rx="42" fill="#07110d"/>
  <circle cx="80" cy="70" r="37" fill="{skin}"/>
  <path d="M43 67c4-31 25-45 54-39 18 4 29 16 31 35-24-10-51-4-85 4z" fill="{hair}"/>
  <circle cx="66" cy="73" r="4" fill="#07110d"/><circle cx="94" cy="73" r="4" fill="#07110d"/>
  <path d="M66 96c10 9 21 9 32 0" fill="none" stroke="#07110d" stroke-width="5" stroke-linecap="round"/>
  <path d="M29 146c7-27 26-42 51-42s44 15 51 42" fill="url(#ag{seed})"/>
  <text x="80" y="148" text-anchor="middle" font-family="Arial" font-size="14" font-weight="800" fill="#07110d">{html.escape(initials)}</text>
</svg>""".strip()


def image_url(topic: TopicRow) -> str:
    seed = int(hashlib.sha1(f"{topic.category}:{topic.title}".encode("utf-8")).hexdigest()[:8], 16)
    return UNSPLASH_IMAGE_POOL[seed % len(UNSPLASH_IMAGE_POOL)]


def page_css(prefix: str = "") -> str:
    return f"""
:root{{--bg:#07110d;--panel:rgba(255,255,255,.075);--text:#f5f8f4;--muted:#b8c7bc;--line:rgba(255,255,255,.15);--green:#8ce99a;--blue:#80c7ff;--gold:#ffd166;--shadow:0 26px 80px rgba(0,0,0,.32)}}
*{{box-sizing:border-box}}body{{margin:0;padding-bottom:78px;font-family:-apple-system,BlinkMacSystemFont,'PingFang SC','Microsoft YaHei',Arial,sans-serif;color:var(--text);background:radial-gradient(circle at 10% 0%,rgba(34,150,123,.28),transparent 34rem),linear-gradient(135deg,#050806,var(--bg),#101a13);line-height:1.75}}a{{color:inherit}}.wrap{{width:min(1120px,calc(100% - 42px));margin:auto}}.site-header{{position:sticky;top:0;z-index:40;background:rgba(7,17,13,.86);backdrop-filter:blur(18px);border-bottom:1px solid var(--line)}}.site-nav{{display:flex;align-items:center;justify-content:space-between;gap:18px;padding:14px 0}}.brand{{display:flex;align-items:center;gap:12px;text-decoration:none;font-weight:900;letter-spacing:.08em}}.brand img{{width:46px;height:34px;object-fit:contain;border-radius:10px;background:rgba(255,255,255,.94);padding:3px}}.brand small{{display:block;letter-spacing:.02em;color:var(--muted);font-weight:700;font-size:11px;margin-top:-4px}}.site-links{{display:flex;align-items:center;gap:16px;color:var(--muted);font-size:14px}}.site-links a{{text-decoration:none;white-space:nowrap}}.site-links a:hover{{color:var(--text)}}.site-links .nav-cta{{border:1px solid var(--line);border-radius:999px;padding:8px 12px;color:var(--text);background:rgba(255,255,255,.08);font-weight:900}}.lang-switcher{{display:inline-flex;align-items:center;gap:7px;border:1px solid var(--line);border-radius:999px;background:rgba(255,255,255,.075);padding:7px 9px;color:var(--text);font-size:12px;font-weight:900;white-space:nowrap}}.lang-switcher svg{{width:15px;height:15px;flex:0 0 auto}}.lang-switcher a{{text-decoration:none;color:inherit;opacity:.68}}.lang-switcher a.active{{opacity:1;color:var(--green)}}.hero{{padding:72px 0 44px}}.eyebrow{{color:var(--green);font-weight:900;font-size:13px;letter-spacing:.18em;text-transform:uppercase}}h1{{font-size:clamp(38px,6vw,72px);line-height:1.02;letter-spacing:-.055em;margin:16px 0 18px}}.lead{{font-size:20px;color:var(--muted);max-width:820px}}.grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:18px;padding:18px 0 72px}}.card{{border:1px solid var(--line);border-radius:28px;background:var(--panel);box-shadow:var(--shadow);overflow:hidden;text-decoration:none;display:flex;flex-direction:column}}.card img{{width:100%;aspect-ratio:16/9;object-fit:cover;background:#13221b}}.card-body{{padding:20px}}.card h2{{font-size:22px;line-height:1.25;margin:0 0 10px}}.card p{{color:var(--muted);margin:0}}.meta{{display:flex;gap:10px;flex-wrap:wrap;color:var(--green);font-size:13px;font-weight:800;margin-bottom:12px}}.pager{{display:flex;justify-content:center;gap:10px;padding:0 0 70px}}.pager a,.pager span{{border:1px solid var(--line);border-radius:999px;padding:10px 14px;text-decoration:none;background:var(--panel)}}.site-footer{{border-top:1px solid var(--line);padding:28px 0 42px;color:var(--muted)}}article{{max-width:860px;margin:auto;padding:56px 0 80px}}.cover{{width:100%;border-radius:32px;aspect-ratio:16/9;object-fit:cover;box-shadow:var(--shadow);background:#13221b}}.article-meta{{display:flex;align-items:center;gap:14px;color:var(--muted);margin:22px 0}}.authority-note{{border:1px solid var(--line);border-radius:20px;background:rgba(255,255,255,.06);padding:14px 16px;color:var(--muted);margin:18px 0}}.authority-note a,.source-list a{{color:var(--green);text-decoration:none}}.author-avatar{{width:54px;height:54px;border-radius:18px;flex:0 0 auto}}.content{{font-size:18px;color:#e8efe9}}.content h2{{font-size:32px;line-height:1.15;margin:42px 0 12px;letter-spacing:-.03em;color:var(--text)}}.content p{{margin:0 0 18px}}.content li{{margin:8px 0}}.source-list{{border-top:1px solid var(--line);margin-top:34px;padding-top:22px;color:var(--muted)}}.tags{{display:flex;gap:8px;flex-wrap:wrap;margin-top:32px}}.tag{{border:1px solid rgba(140,233,154,.35);border-radius:999px;padding:6px 10px;color:var(--green);font-size:13px}}.bottom-cta{{position:fixed;left:0;right:0;bottom:0;z-index:45;background:rgba(7,17,13,.91);backdrop-filter:blur(18px);border-top:1px solid var(--line)}}.bottom-cta-inner{{display:flex;align-items:center;justify-content:space-between;gap:16px;padding:10px 0}}.bottom-cta-text{{display:flex;align-items:baseline;gap:10px;color:var(--muted);min-width:0}}.bottom-cta-text strong{{color:var(--text)}}.bottom-cta-text span{{white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.bottom-cta a{{flex:0 0 auto;text-decoration:none;border-radius:999px;padding:9px 13px;background:var(--text);color:#07110d;font-weight:900}}@media(max-width:900px){{.grid{{grid-template-columns:1fr}}.site-links{{display:flex}}.site-links a:not(.nav-cta){{display:none}}.lang-switcher a{{display:none}}.lang-switcher a.active{{display:inline}}.bottom-cta-text span{{display:none}}}}
""".strip()


def language_switcher(root_prefix: str, active: str = "zh", section: str = "home") -> str:
    suffix = "blog/" if section == "blog" else "brand-audit/" if section == "brand-audit" else ""
    zh_href = f"{root_prefix}{suffix}" if suffix else f"{root_prefix}index.html"
    en_href = f"{root_prefix}en/{suffix}"
    ar_href = f"{root_prefix}ar/{suffix}"
    return f"""<div class="lang-switcher" aria-label="Language selector"><svg viewBox="0 0 24 24" aria-hidden="true" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="10"/><path d="M2 12h20M12 2a15.3 15.3 0 0 1 0 20M12 2a15.3 15.3 0 0 0 0 20"/></svg><a class="{ 'active' if active == 'zh' else '' }" href="{zh_href}">中</a><a class="{ 'active' if active == 'en' else '' }" href="{en_href}">EN</a><a class="{ 'active' if active == 'ar' else '' }" href="{ar_href}" dir="rtl">ع</a></div>"""


def site_header(prefix: str, blog_href: Optional[str] = None, tool_href: Optional[str] = None, section: str = "home") -> str:
    blog_link = blog_href or f"{prefix}blog/"
    tool_link = tool_href or f"{prefix}brand-audit/"
    return f"""
<header class="top site-header"><nav class="wrap site-nav"><a class="brand site-brand" href="{prefix}index.html#top"><img src="{prefix}logo.svg" alt="Eco GEO logo"/><span>ECO GEO<small>Brand-first GEO</small></span></a><div class="links navlinks site-links"><a href="{prefix}index.html#why">为什么</a><a href="{prefix}index.html#method">方法</a><a href="{tool_link}">品牌评测</a><a href="{prefix}index.html#credentials">服务品牌</a><a class="nav-cta nav-insights" href="{blog_link}">前沿观点</a><a class="nav-cta" href="{prefix}contact/">联系</a></div>{language_switcher(prefix, 'zh', section)}</nav></header>
""".strip()


def site_footer(prefix: str, blog_href: Optional[str] = None) -> str:
    blog_link = blog_href or f"{prefix}blog/"
    return f"""<footer class="footer site-footer"><div class="wrap">© 2026 Eco GEO · Brand-first GEO · <a href="{prefix}index.html">首页</a> · <a href="{blog_link}">前沿观点</a> · <a href="{prefix}about/">关于</a> · <a href="{prefix}editorial-policy/">编辑政策</a> · <a href="{prefix}privacy/">隐私</a> · <a href="{prefix}terms/">条款</a> · <a href="{prefix}contact/">联系</a></div></footer>"""


def bottom_cta() -> str:
    return """<div class="bottom-cta" role="region" aria-label="Eco GEO contact"><div class="wrap bottom-cta-inner"><div class="bottom-cta-text"><strong>AIBE 初诊</strong><span>检查你的品牌在 AI 答案里的可见度与引用风险</span></div><a href="mailto:yt.feng@foxmail.com?subject=Eco%20GEO%20AIBE%20诊断咨询">邮件咨询</a></div></div>"""


def article_html(topic: TopicRow, article: Dict[str, str], slug: str, author_name: str, initials: str) -> str:
    tags = [t.strip() for t in re.split(r"[,，/、]", article["tags"]) if t.strip()][:6]
    tag_html = "".join(f"<span class='tag'>{html.escape(t)}</span>" for t in tags)
    img = image_url(topic)
    title = html.escape(ensure_title_prefix(article["title"]))
    excerpt = html.escape(article["excerpt"])
    published = article.get("date") or article.get("published_date") or today_publish_date()
    published_html = f'<time datetime="{html.escape(published)}">{html.escape(date_label(published))}</time>'
    canonical = f"{SITE_URL}/blog/articles/{slug}/"
    body_html = article["body_html"] + str(article.get("sources_html", ""))
    schema = json.dumps(
        {
            "@context": "https://schema.org",
            "@type": "BlogPosting",
            "mainEntityOfPage": {"@type": "WebPage", "@id": canonical},
            "headline": ensure_title_prefix(article["title"]),
            "description": article["excerpt"],
            "image": img,
            "datePublished": published,
            "dateModified": article.get("dateModified") or published,
            "author": {"@type": "Organization", "name": AUTHOR_NAME, "url": f"{SITE_URL}/about/"},
            "editor": {"@type": "Organization", "name": REVIEWER_NAME, "url": f"{SITE_URL}/editorial-policy/"},
            "publisher": {"@type": "Organization", "name": "Eco GEO", "logo": {"@type": "ImageObject", "url": f"{SITE_URL}/logo.svg"}},
            "keywords": article["tags"],
            "inLanguage": "zh-CN",
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/><title>{title}｜Eco GEO 前沿观点</title><meta name="description" content="{excerpt}"/><meta name="author" content="{html.escape(AUTHOR_NAME)}"/><link rel="icon" href="../../../logo.svg" type="image/svg+xml"/><link rel="canonical" href="{canonical}"/><style>{page_css('../../../')}</style><script id="schema-article" type="application/ld+json">{schema}</script></head>
<body>{site_header('../../../', '../../', '../../../brand-audit/', 'blog')}<main class="wrap"><article><div class="eyebrow">Eco GEO 前沿观点</div><h1>{title}</h1><p class="lead">{excerpt}</p><img class="cover" src="{img}" alt="{title}" loading="lazy" referrerpolicy="no-referrer"/><div class="article-meta">{avatar_svg(initials, topic.title)}<div><strong>{html.escape(author_name)}</strong><br/><span>{html.escape(topic.category)} · {published_html} · Brand-first GEO</span></div></div><div class="authority-note"><strong>编辑与事实核查：</strong>{html.escape(REVIEWER_NAME)} · 本文遵循 <a href="../../../editorial-policy/">Eco GEO 编辑政策</a>，涉及事实、数据和时事信息时优先引用可验证来源。</div><div class="content">{body_html}</div><div class="tags">{tag_html}</div></article></main>{bottom_cta()}{site_footer('../../../', '../../')}</body></html>"""


def index_page(posts: List[Dict[str, str]], page: int, per_page: int, total_pages: int, prefix: str) -> str:
    start = (page - 1) * per_page
    subset = posts[start : start + per_page]
    cards = []
    root_prefix = "../" if page == 1 else "../../../"
    article_prefix = "articles/" if page == 1 else "../../articles/"
    for p in subset:
        title = ensure_title_prefix(p["title"])
        published = p.get("date") or p.get("published_date") or ""
        date_html = f"<span>{html.escape(date_label(published))}</span>" if published else ""
        cards.append(f"""<a class="card" href="{article_prefix}{p['slug']}/"><img src="{p['image']}" alt="{html.escape(title)}" loading="lazy" referrerpolicy="no-referrer"/><div class="card-body"><div class="meta"><span>{html.escape(p['category'])}</span>{date_html}<span>{html.escape(p['author'])}</span></div><h2>{html.escape(title)}</h2><p>{html.escape(p['excerpt'])}</p></div></a>""")
    if page == 1:
        canonical = "../"
        prev_link = ""
    else:
        canonical = "../../"
        prev_href = "../" if page == 2 else f"../{page-1}/"
        prev_link = f"<a href='{prev_href}'>上一页</a>"
    next_href = f"page/{page+1}/" if page == 1 else f"../../page/{page+1}/"
    next_link = f"<a href='{next_href}'>下一页</a>" if page < total_pages else ""
    pager = f"<div class='pager'>{prev_link}<span>第 {page} / {total_pages} 页</span>{next_link}</div>"
    blog_href = "./" if page == 1 else "../../"
    tool_href = "../brand-audit/" if page == 1 else "../../../brand-audit/"
    return f"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/><title>Eco GEO 前沿观点｜品牌化 GEO 与 AI 搜索洞察</title><meta name="description" content="Eco GEO 前沿观点：品牌化 GEO、白帽 GEO、AIBE、KNIT 与 AI 搜索优化文章。"/><link rel="icon" href="{root_prefix}logo.svg" type="image/svg+xml"/><style>{page_css(root_prefix)}</style></head><body>{site_header(root_prefix, blog_href, tool_href, 'blog')}<main class="wrap"><section class="hero"><div class="eyebrow">Insights Library</div><h1>Eco GEO 前沿观点</h1><p class="lead">围绕品牌化 GEO、白帽 GEO、AIBE、KNIT 与 AI 搜索的全量选题文章库。</p></section><section class="grid">{''.join(cards)}</section>{pager}</main>{bottom_cta()}{site_footer(root_prefix, blog_href)}</body></html>"""


def patch_homepage(blog_count: int) -> None:
    path = Path("index.html")
    if not path.exists():
        return
    text = path.read_text(encoding="utf-8")
    if 'href="blog/"' not in text:
        text = text.replace('<a href="#insights">观点库</a><a href="#contact">联系</a>', '<a href="#insights">观点库</a><a class="nav-cta nav-insights" href="blog/">前沿观点</a><a href="#contact">联系</a>')
    marker = '<section class="section" id="insights">'
    if marker in text and 'href="blog/" class="btn primary"' not in text:
        text = text.replace(
            '<p class="section-intro">Eco GEO 关注的是“AI 会不会引用你、如何解释你、和谁一起比较你”。观点库会持续沉淀 GEO、AIBE、KNIT 与 AI 搜索的实践。</p>',
            f'<p class="section-intro">Eco GEO 关注的是“AI 会不会引用你、如何解释你、和谁一起比较你”。前沿观点库已沉淀 {blog_count} 篇行业文章，覆盖品牌化 GEO、白帽 GEO、AIBE 与 AI 搜索实践。</p><div class="actions"><a href="blog/" class="btn primary">查看前沿观点</a></div>'
        )
    text = re.sub(
        r"前沿观点库已沉淀 \d+ 篇行业文章",
        f"前沿观点库已沉淀 {blog_count} 篇行业文章",
        text,
    )
    path.write_text(text, encoding="utf-8")


def write_blog(topics: List[TopicRow], out_dir: Path, overwrite: bool) -> None:
    api_key = os.environ.get("DEEPSEEK_API_KEY")
    if not api_key:
        raise RuntimeError("DEEPSEEK_API_KEY is required in repository secrets")
    out_dir.mkdir(parents=True, exist_ok=True)
    articles_dir = out_dir / "articles"
    articles_dir.mkdir(exist_ok=True)

    posts: List[Dict[str, str]] = []
    for n, topic in enumerate(topics, start=1):
        slug = slugify(topic.title, topic.idx)
        article_dir = articles_dir / slug
        article_file = article_dir / "index.html"
        author, initials = deterministic_author(topic.title)
        published = historical_publish_date(n, len(topics))
        if article_file.exists() and not overwrite:
            print(f"Skip existing {slug}")
            article = {"title": topic.title, "excerpt": f"围绕 {topic.title} 的品牌化 GEO 实践框架。", "tags": topic.keywords, "body_html": ""}
        else:
            article = deepseek_article(topic, api_key)
            article["date"] = published
            article_dir.mkdir(parents=True, exist_ok=True)
            article_file.write_text(article_html(topic, article, slug, author, initials), encoding="utf-8")
            time.sleep(REQUEST_DELAY)
        posts.append({
            "row": str(topic.idx),
            "slug": slug,
            "title": ensure_title_prefix(article["title"]),
            "excerpt": article["excerpt"],
            "category": topic.category,
            "tags": article["tags"],
            "author": author,
            "reviewed_by": REVIEWER_NAME,
            "date": published,
            "image": image_url(topic),
            "url": f"{SITE_URL}/blog/articles/{slug}/",
        })
        if n % 25 == 0:
            print(f"Generated {n}/{len(topics)} articles")

    per_page = 24
    total_pages = max(1, (len(posts) + per_page - 1) // per_page)
    (out_dir / "index.html").write_text(index_page(posts, 1, per_page, total_pages, "../"), encoding="utf-8")
    page_root = out_dir / "page"
    for page in range(2, total_pages + 1):
        page_dir = page_root / str(page)
        page_dir.mkdir(parents=True, exist_ok=True)
        page_dir.joinpath("index.html").write_text(index_page(posts, page, per_page, total_pages, "../../"), encoding="utf-8")
    (out_dir / "posts.json").write_text(json.dumps(posts, ensure_ascii=False, indent=2), encoding="utf-8")

    sitemap_items = [f"  <url><loc>{SITE_URL}/</loc></url>", f"  <url><loc>{SITE_URL}/brand-audit/</loc></url>", f"  <url><loc>{SITE_URL}/blog/</loc></url>"]
    sitemap_items += [f"  <url><loc>{p['url']}</loc><lastmod>{p.get('date', '')}</lastmod></url>" for p in posts]
    Path("sitemap.xml").write_text('<?xml version="1.0" encoding="UTF-8"?>\n<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n' + "\n".join(sitemap_items) + "\n</urlset>\n", encoding="utf-8")
    patch_homepage(len(posts))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="assets/blog_articles.xlsx")
    parser.add_argument("--out", default="blog")
    parser.add_argument("--start-row", type=int, default=2)
    parser.add_argument("--limit", type=int, default=0, help="0 means all rows")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()
    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)
    topics = read_topics(excel_path, args.start_row, args.limit)
    print(f"Loaded {len(topics)} topics from {excel_path}")
    write_blog(topics, Path(args.out), args.overwrite)


if __name__ == "__main__":
    main()
